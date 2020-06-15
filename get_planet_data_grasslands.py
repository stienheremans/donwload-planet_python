# -*- coding: utf-8 -*-
import json
from osgeo import ogr
import geopandas as gpd
import matplotlib.pyplot as plt
import geojson
import os
import requests
from requests.auth import HTTPBasicAuth
import shapely
from shapely.geometry import box, Polygon
import pandas as pd
import gdal

# import the shapefile of your study site and transform it into a geojson
driver = ogr.GetDriverByName('ESRI Shapefile')
shp_path = "F:/Postdoc 2020-2024/Projects/Graslandonderzoek/1. Inputs/Study sites bboxes/Turnhout_bbox.shp"
file = gpd.read_file(shp_path)
file.to_file("F:/Postdoc 2020-2024/Projects/Graslandonderzoek/2. Methodology/Part 1 - Creating time series/C - Planet data/GeoJSON/Turnhout_gras.json", driver="GeoJSON")


with open("F:/Postdoc 2020-2024/Projects/Graslandonderzoek/2. Methodology/Part 1 - Creating time series/C - Planet data/GeoJSON/Turnhout_gras.json") as f:
    gj = geojson.load(f)
features = gj['features'][0]
  
geometry = {
        "type": "Polygon",
        "coordinates": [
        [
                []
                
        ]
    ]
}
# replace coordinates in the empty json file created above by the actual coordinates from the geojson file
geometry["coordinates"] = features.geometry.coordinates
print(geometry)



# get images that overlap with your study site
geometry_filter = {
        "type": "GeometryFilter",
        "field_name": "geometry",
        "config":geometry
}

# get images acquired within a preset date range
date_range_filter = {
        "type": "DateRangeFilter",
        "field_name": "acquired",
        "config": {
                "gte": "2018-03-01T00:00:00.000Z",
                "lte": "2018-10-31T00:00:00.000Z"
        }
}

# only get images which have < X% cloud coverage
cloud_cover_filter = {
        "type": "RangeFilter",
        "field_name": "cloud_cover",
        "config": {
                "lte": 0.05
    }
}

# combine the geo, date and cloud filters
combined_filter = {
        "type": "AndFilter",
        "config": [geometry_filter, date_range_filter, cloud_cover_filter]
}

# Give API key for ypur Planet account
PLANET_API_KEY = 'c7e872b7d5054e3e839d74ef1b241655'

item_type = "PSScene4Band"
asset_type = "analytic_sr"

# make API request object
search_request = {
        "interval": "day",
        "item_types":[item_type],
        "filter": combined_filter
}

# Fire off the POST request
search_result= \
requests.post(
        'https://api.planet.com/data/v1/quick-search',
        auth=HTTPBasicAuth(PLANET_API_KEY, ''),
        json=search_request)
print(json.dumps(search_result.json(), indent = 1))

# Extract image properties
image_ids = [feature['id'] for feature in search_result.json()['features']]
image_dates = [feature['properties']['acquired'][0:10] for feature in search_result.json()['features']]
image_months = [feature['properties']['acquired'][5:7] for feature in search_result.json()['features']]
image_days = [feature['properties']['acquired'][8:10] for feature in search_result.json()['features']]
image_clouds = [feature['properties']['cloud_cover'] for feature in search_result.json()['features']]

# Determine overlap between study site and each of the planet scenes
poly_hull = Polygon(geometry["coordinates"][0])
crs = {'init': 'epsg:4326'}
geo_hull = gpd.GeoDataFrame(index=[0], crs=crs, geometry = [poly_hull])  
geo_hull_tr =geo_hull.to_crs({'init': 'epsg:32631'})
hull_area = geo_hull_tr['geometry'].area[0]

df_results = pd.DataFrame(index = range(len(image_ids)), columns=['im_id', 'overlap', 'clouds', 'date', 'month', 'day', 'period', 'deviat', 'anal_sr'])
df_results['im_id']= image_ids
df_results['date']= image_dates
df_results['month']= image_months
df_results['day']= image_days
df_results['clouds']= image_clouds
df_results['day'] = df_results['day'].astype('int32')


for x in range(len(image_ids)):
    anal_sr =  'assets.analytic_sr:download' in  search_result.json()['features'][x]['_permissions']
    df_results['anal_sr'][x] = anal_sr
    poly_im = Polygon(search_result.json()['features'][x]['geometry']['coordinates'][0])
    crs = {'init': 'epsg:4326'}
    geo_im = gpd.GeoDataFrame(index=[0], crs=crs, geometry = [poly_im]) 
    geo_im_tr =geo_im.to_crs({'init': 'epsg:32631'})
    geo_inters = geo_im_tr.intersection(geo_hull_tr)
    inters_area = geo_inters.area[0]
    perc_inters = inters_area/hull_area*100
    df_results['overlap'][x] = perc_inters
    if df_results['day'][x]<6:
        df_results['period'][x]= 1
        df_results['deviat'][x]= abs(df_results['day'][x]-3)
    elif df_results['day'][x]<11:
        df_results['period'][x]= 2
        df_results['deviat'][x]= abs(df_results['day'][x]-8)
    elif df_results['day'][x]<16:
        df_results['period'][x]= 3
        df_results['deviat'][x]= abs(df_results['day'][x]-13)
    elif df_results['day'][x]<21:
        df_results['period'][x]= 4
        df_results['deviat'][x]= abs(df_results['day'][x]-18)
    elif df_results['day'][x]<26:
        df_results['period'][x]= 5
        df_results['deviat'][x]= abs(df_results['day'][x]-23)
    else: 
        df_results['period'][x]= 6
        df_results['deviat'][x]= abs(df_results['day'][x]-28)
    

df_results['overlap'] = df_results['overlap'].astype('float')

# Select only the images with more than 80% overlap with the study area    
df_results2 = df_results[df_results['overlap']>=80]

# Select only the images with downloadable analytic_sr asset
df_results2 = df_results2[df_results2['anal_sr']==True]

# Select only the images with minimum clouds per period
def func(group):
    return group.loc[group['overlap'] == group['overlap'].max()]

df_results3 = df_results2.groupby(['month','period'], as_index=False).apply(func).reset_index(drop=True)

# Select the image with minimal clouds that is closest to the mid of the period
def func2(group):
    return group.loc[group['deviat'] == group['deviat'].min()]

df_results4 = df_results3.groupby(['month','period'], as_index=False).apply(func2).reset_index(drop=True)

# Write all df_results to an excel file (in different tabs)
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

filename = "F:/Postdoc 2020-2024/Projects/Graslandonderzoek/2. Methodology/Part 1 - Creating time series/C - Planet data/Image lists/Planet_imgs_Turnhout_2018.xlsx"
wb = Workbook()
sheet1 = wb.create_sheet('all_imgs',0)
sheet2 = wb.create_sheet('imgs_overlap_anal',1)
sheet3 = wb.create_sheet('imgs_min_clouds',2)
sheet4 = wb.create_sheet('imgs_min_deviat',3)

for x in dataframe_to_rows(df_results):
    sheet1.append(x)
    
for x in dataframe_to_rows(df_results2):
    sheet2.append(x)
    
for x in dataframe_to_rows(df_results3):
    sheet3.append(x)
    
for x in dataframe_to_rows(df_results4):
    sheet4.append(x)

wb.save(filename)

    
    